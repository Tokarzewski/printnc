"""
Buduje BOM z osadzonymi zdjeciami elementow (po jednym miniaturze na pozycje).

Laczy dane z BOM/BOM.xlsx ze zdjeciami z folderu, dopasowujac je po 3-cyfrowym ID
(np. plik '006...jpg' trafia do wiersza pozycji ID 6 — zgodnie z Issue #1).
Wynik to NOWY plik xlsx (nie modyfikuje BOM.xlsx).

Uzycie:
    py skrypty/bom_ze_zdjeciami.py [folder_ze_zdjeciami] [plik_wyjsciowy.xlsx]

Domyslnie:
    folder zdjec  = BOM/zdjecia
    plik wyjsciowy = BOM/generated/BOM_ze_zdjeciami.xlsx

Wymaga: openpyxl ORAZ Pillow (openpyxl potrzebuje Pillow do osadzania obrazow).
    py -m pip install openpyxl pillow
"""
import io
import os
import re
import sys

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from bom_common import load_bom_rows
from enrich_bom import categorize

IMG_EXT = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
THUMB_PX = 110  # docelowa szerokosc/wysokosc miniatury w pikselach

COLUMNS = [
    ("ID", 8),
    ("Zdjęcie", 18),
    ("Nazwa", 42),
    ("Ilość", 8),
    ("Wymiary (mm)", 18),
    ("Rodzaj zakupu", 22),
    ("Link Allegro", 50),
]


def index_photos(folder):
    """Mapuje 3-cyfrowe ID -> sciezka pliku zdjecia (pierwszy pasujacy)."""
    found = {}
    if not folder or not os.path.isdir(folder):
        return found
    for f in sorted(os.listdir(folder)):
        if os.path.splitext(f)[1].lower() not in IMG_EXT:
            continue
        m = re.match(r"(\d{3})", f)
        if m:
            found.setdefault(m.group(1), os.path.join(folder, f))
    return found


def make_thumb(path):
    """Robi miniaturę (Pillow) i zwraca (BytesIO, w_px, h_px)."""
    from PIL import Image as PILImage

    im = PILImage.open(path)
    im.thumbnail((THUMB_PX, THUMB_PX))
    w, h = im.size
    buf = io.BytesIO()
    fmt = "PNG" if im.mode in ("RGBA", "P", "LA") else "JPEG"
    im.convert("RGBA" if fmt == "PNG" else "RGB").save(buf, fmt)
    buf.seek(0)
    return buf, w, h


def main():
    sys.stdout.reconfigure(encoding="utf-8")
    try:
        import PIL  # noqa: F401
    except ImportError:
        print("Brak Pillow (openpyxl wymaga go do osadzania obrazów). Zainstaluj:\n    py -m pip install pillow")
        sys.exit(2)
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    folder = args[0] if len(args) > 0 else os.path.join("BOM", "zdjecia")
    out = args[1] if len(args) > 1 else os.path.join("BOM", "generated", "BOM_ze_zdjeciami.xlsx")

    rows = load_bom_rows()
    photos = index_photos(folder)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    # naglowek
    for ci, (title, width) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=ci, value=title)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.freeze_panes = "A2"

    embedded = 0
    for ri, r in enumerate(rows, start=2):
        try:
            pid = f"{int(r['id']):03d}"
        except (TypeError, ValueError):
            pid = str(r["id"])
        ws.cell(row=ri, column=1, value=r["id"])
        ws.cell(row=ri, column=3, value=r["nazwa"])
        ws.cell(row=ri, column=4, value=r["ilosc"])
        ws.cell(row=ri, column=5, value=r["wymiary"])
        ws.cell(row=ri, column=6, value=categorize(r))
        ws.cell(row=ri, column=7, value=r["link"])
        for col in (3, 5, 6, 7):
            ws.cell(row=ri, column=col).alignment = Alignment(vertical="center", wrap_text=(col == 3))

        path = photos.get(pid)
        if path:
            data, w, h = make_thumb(path)
            img = XLImage(data)
            img.width, img.height = w, h
            ws.row_dimensions[ri].height = max(h * 0.78, 84)  # punkty ~ piksele*0.75
            ws.add_image(img, f"B{ri}")
            embedded += 1
        else:
            ws.row_dimensions[ri].height = 20

    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"Pozycji: {len(rows)} | zdjęć w folderze '{folder}': {len(photos)} | osadzonych: {embedded}")
    if embedded == 0:
        print("UWAGA: nie osadzono żadnego zdjęcia.")
        print(f"  Wrzuć zdjęcia do '{folder}' z nazwami zaczynającymi się od 3-cyfrowego ID")
        print("  (patrz BOM/generated/01_zdjecia_nazewnictwo.csv), potem uruchom ponownie.")
    print("Zapisano:", out)


if __name__ == "__main__":
    main()
